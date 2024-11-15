# main.py
import cv2
import time
import configparser
from ultralytics import YOLO
from camera_setup import initialize_cameras
from sort import Sort
import threading
import openpyxl
from openpyxl import Workbook
import os
import numpy as np
import zipfile
import vehicle_id_manager

def setup_model():
    model = YOLO("yolov5nu.pt")
    return model

def detect_objects(model, frame, conf_threshold):
    results = model(frame)
    detections = []
    for result in results:
        boxes = result.boxes
        for box in boxes:
            confidence = box.conf.item()
            if confidence < conf_threshold:
                continue  # Skip detections with low confidence
            x1, y1, x2, y2 = box.xyxy[0].cpu().numpy()
            class_id = int(box.cls.item())
            detections.append([x1, y1, x2, y2, confidence, class_id])
    return detections

def save_to_excel(data, excel_file):
    try:
        if os.path.exists(excel_file):
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(['Timestamp', 'Total Vehicle Count', 'Direction Counts', '', 'Total Vehicles per Hour', 'Unique Vehicles per Hour'])
    except (FileNotFoundError, zipfile.BadZipFile, openpyxl.utils.exceptions.InvalidFileException):
        print(f"Excel file not found or invalid. Creating a new file: {excel_file}")
        wb = Workbook()
        ws = wb.active
        ws.append(['Timestamp', 'Total Vehicle Count', 'Direction Counts', '', 'Total Vehicles per Hour', 'Unique Vehicles per Hour'])
    except Exception as e:
        print(f"Error loading workbook: {e}")
        return
    ws.append(data)
    wb.save(excel_file)

def process_cameras():
    config = configparser.ConfigParser()
    config.read('settings.ini')
    conf_threshold = float(config['Detection']['conf_threshold'])
    excel_file = config['Export']['excel_file']

    # SORT tracker parameters
    max_age = int(config['Tracking']['max_age'])
    min_hits = int(config['Tracking']['min_hits'])
    iou_threshold = float(config['Tracking']['iou_threshold'])

    # Read update interval from settings or default to 10 seconds
    update_interval = int(config.get('Update', 'update_interval', fallback=10))

    caps = initialize_cameras()
    model = setup_model()

    # Initialize a tracker for each camera
    trackers = {}
    for index, _ in caps:
        trackers[index] = Sort(max_age=max_age, min_hits=min_hits, iou_threshold=iou_threshold)

    # Load existing vehicle IDs
    tracked_vehicle_ids, vehicle_id_timestamps = vehicle_id_manager.load_vehicle_ids('vehicle_ids.json')
    print(f"Starting with {len(tracked_vehicle_ids)} tracked vehicle IDs.")

    DESIRED_FPS = int(config['Camera']['fps'])
    frame_interval = 1 / DESIRED_FPS
    last_frame_time = 0

    # Load class names for the COCO dataset
    class_names = model.names

    # For counts over the past hour
    interval_counts = []  # List of (timestamp, vehicle_directions)

    # Timer for updating every update_interval seconds
    last_update_time = time.time()

    # Dictionary to keep track of vehicle IDs and their last seen camera
    vehicle_last_seen = {}  # {vehicle_id: (camera_index, timestamp)}

    try:
        while True:
            current_time = time.time()
            if current_time - last_frame_time < frame_interval:
                time.sleep(0.01)
                continue
            last_frame_time = current_time

            # Process frames from all cameras
            frames = {}
            for index, cap in caps:
                ret, frame = cap.read()
                if not ret:
                    print(f"Failed to read from Camera {index}")
                    continue
                frame_resized = cv2.resize(frame, (640, 480))
                frames[index] = frame_resized

            # Skip processing if no frames were read
            if not frames:
                print("No frames to process.")
                continue

            vehicle_directions = {}  # {vehicle_id: [list of camera indices in order]}

            # Detect and track vehicles in each camera
            current_vehicle_ids = {}
            for index, frame in frames.items():
                # Detect objects
                detections = detect_objects(model, frame, conf_threshold)
                # Prepare detections for SORT
                sort_detections = []
                for det in detections:
                    x1, y1, x2, y2, confidence, class_id = det
                    class_name = class_names[class_id]
                    if class_name in ['car', 'truck']:
                        sort_detections.append([x1, y1, x2, y2, confidence])

                sort_detections = np.array(sort_detections) if sort_detections else np.empty((0, 5))
                # Update tracker
                tracks = trackers[index].update(sort_detections)
                current_vehicle_ids[index] = set()
                for track in tracks:
                    x1, y1, x2, y2, track_id = track
                    track_id = int(track_id)
                    x1, y1, x2, y2 = int(x1), int(y1), int(x2), int(y2)
                    label = f"ID {track_id}"
                    color = (0, 255, 0)
                    cv2.rectangle(frame, (x1, y1), (x2, y2), color, 2)
                    cv2.putText(frame, label, (x1, y1 - 10),
                                cv2.FONT_HERSHEY_SIMPLEX, 0.5, color, 2)
                    # Update vehicle last seen info
                    if track_id not in vehicle_last_seen:
                        vehicle_last_seen[track_id] = []
                    vehicle_last_seen[track_id].append((index, current_time))
                    current_vehicle_ids[index].add(track_id)
                # Display the frame
                cv2.imshow(f"Camera {index}", frame)

            # Determine directions based on the sequence of cameras the vehicle was seen in
            for vehicle_id, sightings in vehicle_last_seen.items():
                # If the vehicle has been seen in more than one camera
                if len(sightings) > 1:
                    # Sort sightings by timestamp
                    sightings.sort(key=lambda x: x[1])
                    # Extract camera indices
                    camera_sequence = [s[0] for s in sightings]
                    vehicle_directions[vehicle_id] = camera_sequence

            # Update counts
            interval_counts.append((current_time, vehicle_directions.copy()))
            # Remove old entries beyond one hour
            interval_counts = [(t, dirs) for t, dirs in interval_counts if current_time - t <= 3600]

            # Update and save to Excel every update_interval seconds
            if current_time - last_update_time >= update_interval:
                # Compute total vehicles per hour (unique vehicle IDs seen in the past hour)
                total_vehicle_ids = set()
                for _, dirs in interval_counts:
                    total_vehicle_ids.update(dirs.keys())
                total_vehicles_per_hour = len(total_vehicle_ids)
                # Compute unique vehicles per hour (vehicle IDs currently being tracked)
                unique_vehicles_per_hour = len(vehicle_last_seen)
                # For current update, count vehicles moving between cameras
                direction_counts = {}
                for dirs in vehicle_directions.values():
                    direction = ' -> '.join(map(str, dirs))
                    direction_counts[direction] = direction_counts.get(direction, 0) + 1
                # Prepare direction counts as a string
                direction_counts_str = ', '.join([f"{k}: {v}" for k, v in direction_counts.items()])
                # Total vehicle count
                total_vehicle_count = sum(len(ids) for ids in current_vehicle_ids.values())
                # Prepare data for Excel
                timestamp_str = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(current_time))
                data = [
                    timestamp_str,
                    total_vehicle_count,
                    direction_counts_str,
                    '',
                    total_vehicles_per_hour,
                    unique_vehicles_per_hour
                ]
                save_to_excel(data, excel_file)
                print(f"Data saved to {excel_file} at {timestamp_str}")
                last_update_time = current_time  # Reset last update time

            if cv2.waitKey(1) & 0xFF == ord('q'):
                break

    finally:
        # Save vehicle IDs when the program is closed
        vehicle_id_manager.save_vehicle_ids(tracked_vehicle_ids, vehicle_id_timestamps, 'vehicle_ids.json')
        for _, cap in caps:
            cap.release()
        cv2.destroyAllWindows()

if __name__ == "__main__":
    process_cameras()

